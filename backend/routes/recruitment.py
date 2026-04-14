from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Response, Form, BackgroundTasks
from fastapi.responses import StreamingResponse, RedirectResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import pandas as pd
from io import BytesIO
import json
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import logging
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from fastapi import BackgroundTasks
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

from database import get_db
from models import Candidate, RecruitmentStats, CandidateStatus, User
from schemas.recruitment import (
    CandidateCreate,
    CandidateUpdate,
    CandidateResponse,
    RecruitmentStatsResponse
)
from routes.auth import get_current_user
from pydantic import ValidationError

router = APIRouter(prefix="/api/recruitment", tags=["recruitment"])

@router.get("/candidates", response_model=List[CandidateResponse])
async def get_candidates(
    stage: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Candidate)

    # Filter by company name for admin and employees from the same company
    if current_user.company_name:
        # Check if the user is an admin OR an employee belonging to the same company as an admin
        # Assuming employees have a company_name field populated if they are associated with a company
        # and that regular employees should only see candidates if their company matches the admin's company
        # If the requirement is that ANY employee can see candidates from their company, the logic would be slightly different.
        # Based on the request "The employees from the same compay should be able to view the candiates same as the company admin"
        # I'm interpreting this as employees within a company should see that company's candidates.

        # Option 1: Admin sees all candidates in their company, employees see candidates in their company
        # This assumes all relevant employees have company_name set.
        query = query.filter(Candidate.company_name == current_user.company_name)

        # Option 2: Only admins and employees specifically marked can access recruitment, and they see their company's candidates
        # This aligns with the existing access control in Recruitment.tsx
        # if current_user.role == "admin" or (current_user.role == "employee" and current_user.employee and current_user.employee.can_access_recruitment):
        #     query = query.filter(Candidate.company_name == current_user.company_name)
        # I will go with a simplified version of Option 1 for now, assuming company_name is the primary filter.

    if stage:
        query = query.filter(Candidate.stage == stage.lower())
    if status:
        query = query.filter(Candidate.status == status.lower())

    # Execute the query and fetch candidates
    candidates = query.all()

    # Manually add the creator_name to each candidate response
    candidate_responses = []
    for candidate in candidates:
        candidate_data = candidate.__dict__
        if candidate.creator:
            candidate_data["creator_name"] = candidate.creator.name
        else:
            candidate_data["creator_name"] = "N/A" # Or some default value
        candidate_responses.append(CandidateResponse(**candidate_data))

    return candidate_responses

@router.post("/candidates", response_model=CandidateResponse)
async def create_candidate(
    candidate: CandidateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    try:
        # Convert the candidate data to a dictionary and exclude None values
        candidate_data = candidate.model_dump(exclude_unset=True)
        
        # Set company_name from current user if available
        if current_user.company_name:
            candidate_data['company_name'] = current_user.company_name
        
        # Ensure stage and status are lowercase
        if 'stage' in candidate_data:
            candidate_data['stage'] = candidate_data['stage'].lower()
        if 'status' in candidate_data:
            candidate_data['status'] = candidate_data['status'].lower()
        
        # Create a new candidate instance
        db_candidate = Candidate(**candidate_data)
        
        # Set the creator user ID
        db_candidate.created_by_user_id = current_user.id

        # Add and commit to database
        db.add(db_candidate)
        db.commit()
        db.refresh(db_candidate)
        
        return db_candidate
    except Exception as e:
        db.rollback()
        logging.error(f"Error creating candidate: {str(e)}")
        if "column \"notes\" of relation \"candidates\" does not exist" in str(e):
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database schema is out of date. Please run database migrations."
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating candidate: {str(e)}"
        )

@router.get("/candidates/{candidate_id}", response_model=CandidateResponse)
async def get_candidate(
    candidate_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    return candidate

@router.put("/candidates/{candidate_id}", response_model=CandidateResponse)
async def update_candidate(
    candidate_id: int,
    candidate: CandidateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not db_candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    # Update candidate fields
    for field, value in candidate.model_dump(exclude_unset=True).items():
        setattr(db_candidate, field, value)
    
    try:
        db.commit()
        db.refresh(db_candidate)
        return db_candidate
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating candidate: {str(e)}"
        )

@router.delete("/candidates/{candidate_id}")
async def delete_candidate(
    candidate_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    # Delete the resume file if it exists
    if candidate.resume_url and os.path.exists(candidate.resume_url):
        try:
            os.remove(candidate.resume_url)
        except Exception as e:
            print(f"Error deleting resume file: {str(e)}")
    
    # Delete the candidate from the database
    db.delete(candidate)
    db.commit()
    
    return {"message": "Candidate deleted successfully"}

@router.get("/template")
async def get_template(current_user: User = Depends(get_current_user)):
    try:
        # Create a DataFrame with the template structure
        df = pd.DataFrame(columns=[
            'name', 'email', 'phone', 'position', 'experience',
            'skills', 'notes', 'stage', 'status'
        ])
        
        # Add a sample row with instructions
        sample_data = {
            'name': 'John Doe',
            'email': 'john.doe@example.com',
            'phone': '+1234567890',
            'position': 'Software Engineer',
            'experience': '5',
            'skills': 'Python, FastAPI, React',
            'notes': 'Sample candidate notes',
            'stage': 'applied',
            'status': 'active'
        }
        df = pd.concat([df, pd.DataFrame([sample_data])], ignore_index=True)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Candidates')
            
            # Get the worksheet
            worksheet = writer.sheets['Candidates']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Prepare the response
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=candidate_template.xlsx"
            }
        )
    except Exception as e:
        logging.error(f"Error generating template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating template: {str(e)}"
        )

@router.post("/bulk-upload")
async def bulk_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    try:
        # Read the Excel file
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))

        # Validate required columns
        required_columns = ['name', 'email', 'position']
        missing_columns = [col for col in df.columns if col.lower() in [rc.lower() for rc in required_columns] and df[col].isnull().all()]

        if missing_columns:
             raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing data in required columns: {', '.join(missing_columns)}"
             )

        # Process each row
        candidates = []
        for index, row in df.iterrows():
            try:
                # Convert row to dictionary and handle NaN values
                candidate_data = row.to_dict()
                candidate_data = {k: str(v) if pd.notna(v) else None for k, v in candidate_data.items()}

                # Map column names to schema attributes (case-insensitive)
                mapped_data = {}
                schema_fields = ['name', 'email', 'phone', 'position', 'experience', 'skills', 'notes', 'stage', 'status']
                for field in schema_fields:
                    # Find the column in df.columns that matches schema field case-insensitively
                    matching_col = next((col for col in df.columns if col.lower() == field.lower()), None)
                    if matching_col and candidate_data.get(matching_col) is not None:
                         mapped_data[field] = candidate_data[matching_col]

                # Set company_name from current user if they have one
                if current_user.company_name:
                    mapped_data['company_name'] = current_user.company_name

                # Set default values for optional fields if not provided or empty
                if 'stage' not in mapped_data or not mapped_data['stage']:
                    mapped_data['stage'] = 'applied' # Default stage
                # Ensure stage is lowercase to match enum
                mapped_data['stage'] = mapped_data['stage'].lower()

                if 'status' not in mapped_data or not mapped_data['status']:
                    mapped_data['status'] = 'active' # Default status
                # Ensure status is lowercase
                mapped_data['status'] = mapped_data['status'].lower()

                # Handle experience conversion
                if 'experience' in mapped_data and mapped_data['experience'] is not None:
                     try:
                         mapped_data['experience'] = float(mapped_data['experience'])
                     except (ValueError, TypeError):
                         mapped_data['experience'] = None # Set to None if conversion fails

                # Handle skills (ensure it's a string, join list if necessary)
                if 'skills' in mapped_data and mapped_data['skills'] is not None:
                    if isinstance(mapped_data['skills'], list):
                         mapped_data['skills'] = ','.join(str(s).strip() for s in mapped_data['skills'] if s)
                    else:
                         mapped_data['skills'] = str(mapped_data['skills']).strip()
                else:
                     mapped_data['skills'] = ""

                # Remove any keys that are not part of the schema or have None values that should be defaulted
                final_candidate_data = {
                     k: v for k, v in mapped_data.items() if k in schema_fields or k == 'company_name'
                }

                # Validate against the schema (optional but good practice)
                # from schemas.recruitment import CandidateCreate
                # CandidateCreate(**final_candidate_data)

                # Create candidate object
                candidate = Candidate(**final_candidate_data)
                
                # Set the creator user ID for bulk uploaded candidates
                candidate.created_by_user_id = current_user.id

                candidates.append(candidate)

            except Exception as e:
                logging.error(f"Error processing row {index + 2}: {str(e)}") # +2 for 0-index and header
                # Decide how to handle errors: skip row, collect errors, etc.
                # For now, we'll just log and skip
                continue # Skip this row and continue with the next

        if not candidates:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid candidates found in the file or an error occurred processing all rows."
            )

        # Bulk insert candidates
        try:
            db.add_all(candidates)
            db.commit()
            return {"message": f"Successfully uploaded {len(candidates)} candidates"}
        except Exception as e:
            db.rollback()
            # Check for common database errors, e.g., missing columns due to migrations
            if "column " in str(e).lower() and " does not exist" in str(e).lower():
                 raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Database schema is out of date. Please run database migrations. " + str(e)
                 )
            # Check for duplicate email errors
            if "duplicate key value violates unique constraint" in str(e).lower() and "candidates_email_key" in str(e).lower():
                 raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="One or more candidates with the same email already exist."
                 )
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error saving candidates to database: {str(e)}"
            )

    except HTTPException as e:
        # Re-raise HTTPException to be handled by FastAPI
        raise e
    except Exception as e:
        logging.error(f"Error processing bulk upload file: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )

@router.get("/stats", response_model=RecruitmentStatsResponse)
async def get_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get recruitment statistics"""
    try:
        # Start with the base query for candidates
        query = db.query(Candidate)

        # Filter by company name for users associated with a company
        if current_user.company_name:
            query = query.filter(Candidate.company_name == current_user.company_name)

        # Get total candidates
        total_candidates = query.count() # Use count() on the filtered query

        # Get active candidates (in screening, interview, or offer stages)
        active_candidates = query.filter(
            Candidate.stage.in_(["screening", "interview", "offer"])
        ).count() # Use count() on the filtered query

        # Get hired candidates
        hired_candidates = query.filter(
            Candidate.stage == "hired"
        ).count() # Use count() on the filtered query

        # Get applied candidates
        applied_candidates = query.filter(
            Candidate.stage == "applied"
        ).count() # Use count() on the filtered query

        return {
            "total_candidates": total_candidates,
            "active_candidates": active_candidates,
            "hired_candidates": hired_candidates,
            "applied_candidates": applied_candidates
        }
    except Exception as e:
        logging.error(f"Error fetching recruitment stats: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching recruitment stats: {str(e)}"
        )

def update_recruitment_stats(db: Session):
    """Update recruitment statistics"""
    try:
        stats = db.query(RecruitmentStats).first()
        if not stats:
            stats = RecruitmentStats()
            db.add(stats)
        
        # Get counts from database
        total_candidates = db.query(Candidate).count()
        active_candidates = db.query(Candidate).filter(
            Candidate.stage.in_(["screening", "interview", "offer"])
        ).count()
        hired_candidates = db.query(Candidate).filter(
            Candidate.stage == "hired"
        ).count()
        applied_candidates = db.query(Candidate).filter(
            Candidate.stage == "applied"
        ).count()
        
        # Update stats
        stats.total_candidates = total_candidates
        stats.active_candidates = active_candidates
        stats.hired_candidates = hired_candidates
        stats.applied_candidates = applied_candidates
        stats.updated_at = datetime.utcnow()
        
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Error updating recruitment stats: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating recruitment stats: {str(e)}"
        )

@router.put("/candidates/{candidate_id}/resume", response_model=CandidateResponse)
async def update_candidate_resume(
    candidate_id: int,
    resume: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not db_candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    # Save the resume file
    try:
        # Create uploads directory if it doesn't exist
        os.makedirs("uploads/resumes", exist_ok=True)
        
        # Generate a unique filename
        file_extension = os.path.splitext(resume.filename)[1]
        unique_filename = f"{candidate_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_extension}"
        file_path = f"uploads/resumes/{unique_filename}"
        
        # Save the file
        with open(file_path, "wb+") as file_object:
            file_object.write(resume.file.read())
        
        # Update the resume_url in database
        db_candidate.resume_url = file_path
        db.commit()
        db.refresh(db_candidate)
        
        return db_candidate
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error uploading resume: {str(e)}"
        )

@router.get("/candidates/{candidate_id}/view-profile")
async def view_candidate_profile(
    candidate_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    if not candidate.resume_url:
        raise HTTPException(status_code=404, detail="No resume found for this candidate")
    
    try:
        # Read the file
        with open(candidate.resume_url, "rb") as file:
            file_content = file.read()
        
        # Return the file as a streaming response
        return StreamingResponse(
            io.BytesIO(file_content),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename=resume_{candidate_id}.pdf"
            }
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Resume file not found")
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error reading resume file: {str(e)}"
        )

async def send_email_async(recipient_email: str, subject: str, body: str):
    try:
        # Validate email configuration
        email_user = os.getenv('EMAIL_HOST_USER')
        email_password = os.getenv('EMAIL_HOST_PASSWORD')
        email_host = os.getenv('EMAIL_HOST', 'smtp.gmail.com')
        email_port = int(os.getenv('EMAIL_PORT', '587'))

        if not email_user or not email_password:
            logging.error("Email credentials not configured")
            raise Exception("Email credentials not properly configured")

        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'plain'))
        
        logging.info(f"Attempting to send email to {recipient_email} via {email_host}:{email_port}")
        
        # Create SMTP session
        server = smtplib.SMTP(email_host, email_port)
        server.starttls()
        
        logging.info("Connected to SMTP server, attempting login...")
        server.login(email_user, email_password)
        
        logging.info("Login successful, sending email...")
        text = msg.as_string()
        server.sendmail(email_user, recipient_email, text)
        server.quit()
        
        logging.info(f"Email sent successfully to {recipient_email}")
    except smtplib.SMTPAuthenticationError as e:
        logging.error(f"SMTP Authentication Error: {str(e)}")
        raise Exception("Failed to authenticate with email server. Please check credentials.")
    except smtplib.SMTPException as e:
        logging.error(f"SMTP Error: {str(e)}")
        raise Exception(f"SMTP Error: {str(e)}")
    except Exception as e:
        logging.error(f"Error sending email: {str(e)}")
        raise e

@router.post("/candidates/{candidate_id}/send-email")
async def send_candidate_email(
    candidate_id: int,
    background_tasks: BackgroundTasks,
    subject: str = Form(...),
    body: str = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Send email to a candidate"""
    try:
        # Validate input
        if not subject or not subject.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Subject is required"
            )
        if not body or not body.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Body is required"
            )

        candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if not candidate:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Candidate not found"
            )
            
        if not candidate.email:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Candidate has no email address"
            )

        # Validate email configuration
        email_user = os.getenv('EMAIL_HOST_USER')
        email_password = os.getenv('EMAIL_HOST_PASSWORD')
        
        if not email_user or not email_password or email_user == "your_email@gmail.com":
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Email service not properly configured"
            )

        # Add email task to background tasks
        background_tasks.add_task(
            send_email_async,
            recipient_email=candidate.email,
            subject=subject.strip(),
            body=body.strip()
        )

        return {"status": "success", "message": "Email scheduled to be sent"}
    except HTTPException as e:
        raise e
    except Exception as e:
        logging.error(f"Error in send_candidate_email: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )
